# Блок 1: Импорт библиотек
# pandas для данных, re для очистки, os для папок, openpyxl для Excel, reportlab для PDF.
import pandas as pd
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Блок 2: Чтение и очистка первичного файла
# Читаем исходный файл и создаём чистую копию.
try:
    df_raw = pd.read_csv('/home/mangust1981/Документы/3Пайтон/wildberries_data.csv',
                         sep=';', on_bad_lines='skip')
except Exception as e:
    print(f"Ошибка чтения исходного файла: {e}")
    exit()

def clean_number(value):
    if not isinstance(value, str):
        return value
    cleaned = re.sub(r'[^0-9.]', '', value)
    return float(cleaned) if cleaned else 0.0

df_clean = df_raw.copy()
if 'Цена' in df_clean.columns:
    df_clean['Цена'] = df_clean['Цена'].apply(clean_number)
if 'Рейтинг' in df_clean.columns:
    df_clean['Рейтинг'] = df_clean['Рейтинг'].apply(clean_number)

# Создаём папки для результатов.
base_dir = '/home/mangust1981/Документы/3Пайтон/wildberries_data_cleaner'
csv_dir = f'{base_dir}/CSV'
excel_dir = f'{base_dir}/Excel'
pdf_dir = f'{base_dir}/PDF'
os.makedirs(csv_dir, exist_ok=True)
os.makedirs(excel_dir, exist_ok=True)
os.makedirs(pdf_dir, exist_ok=True)

# Сохраняем чистый CSV.
df_clean.to_csv(f'{csv_dir}/Чистые_данные.csv', index=False, sep=',')

# Функция для сохранения Excel с адаптивной шириной и выравниванием.
def save_to_excel(df, filepath):
    wb = Workbook()
    ws = wb.active
    for row in df.itertuples(index=False):
        ws.append([row.Название, row.Цена, row.Рейтинг, row.Ссылка])
    for col in ['B', 'C']:  # Цена (B) и Рейтинг (C).
        for cell in ws[col]:
            cell.alignment = Alignment(horizontal='center')
    max_length = max(len(str(row.Название)) for row in df.itertuples(index=False))
    ws.column_dimensions['A'].width = max_length * 1.2
    wb.save(filepath)

save_to_excel(df_clean, f'{excel_dir}/Чистые_данные.xlsx')

# Блок 3: Функция сохранения в PDF
# Регистрируем шрифт и включаем перенос текста.
pdfmetrics.registerFont(TTFont('DejaVuSans',
                               '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
styles = getSampleStyleSheet()
style = styles['Normal']
style.fontName = 'DejaVuSans'
style.fontSize = 10

def save_to_pdf(df, filepath):
    doc = SimpleDocTemplate(filepath, pagesize=A4)
    # Преобразуем данные в список с переносом текста для Название и Ссылка.
    data = [df.columns.tolist()]  # Заголовки.
    for row in df.itertuples(index=False):
        data.append([
            Paragraph(str(row.Название), style),  # Перенос для Название.
            str(row.Цена),                        # Цена без переноса.
            str(row.Рейтинг),                     # Рейтинг без переноса.
            Paragraph(str(row.Ссылка), style)     # Перенос для Ссылка.
        ])
    table = Table(data, colWidths=[200, 50, 50, 200], rowHeights=30)  # Высота строк увеличена.
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'DejaVuSans'),  # Кириллица.
        ('ALIGN', (1, 1), (2, -1), 'CENTER'),          # Цена и Рейтинг по центру.
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),        # Вертикальное выравнивание.
        ('GRID', (0, 0), (-1, -1), 1, colors.black),   # Сетка.
    ]))
    doc.build([table])

save_to_pdf(df_clean, f'{pdf_dir}/Чистые_данные.pdf')

# Блок 4: Сортировка по цене (от меньшего к большему)
# Сохраняем во всех форматах.
price_asc = df_clean.sort_values(by='Цена', ascending=True, ignore_index=True)
price_asc.to_csv(f'{csv_dir}/Сортировка_по_цене.csv', index=False, sep=',')
save_to_excel(price_asc, f'{excel_dir}/Сортировка_по_цене.xlsx')
save_to_pdf(price_asc, f'{pdf_dir}/Сортировка_по_цене.pdf')

# Блок 5: Сортировка по рейтингу с учётом цены
# Сохраняем с выравниванием и шириной.
rating_asc = df_clean.sort_values(by=['Рейтинг', 'Цена'],
                                  ascending=[True, True],
                                  ignore_index=True)
rating_asc.to_csv(f'{csv_dir}/Сортировка_по_рейтингу.csv', index=False, sep=',')
save_to_excel(rating_asc, f'{excel_dir}/Сортировка_по_рейтингу.xlsx')
save_to_pdf(rating_asc, f'{pdf_dir}/Сортировка_по_рейтингу.pdf')

# Блок 6: Разделение по уникальным рейтингам
# Файлы для каждого рейтинга.
if 'Рейтинг' in df_clean.columns:
    unique_ratings = df_clean['Рейтинг'].dropna().unique()
    for rating in unique_ratings:
        rating_df = df_clean[df_clean['Рейтинг'] == rating]
        rating_df = rating_df.sort_values(by='Цена', ascending=True,
                                          ignore_index=True)
        rating_filename = f'Рейтинг_{rating}'
        rating_df.to_csv(f'{csv_dir}/{rating_filename}.csv',
                         index=False, sep=',')
        save_to_excel(rating_df, f'{excel_dir}/{rating_filename}.xlsx')
        save_to_pdf(rating_df, f'{pdf_dir}/{rating_filename}.pdf')

# Блок 7: Сортировка по ценовым диапазонам
# Диапазоны с выравниванием и шириной.
if 'Цена' in df_clean.columns:
    price_ranges = [
        (0, 1000, 'Цена_0_до_1000'),
        (1001, 2000, 'Цена_1001_до_2000'),
        (2001, 3000, 'Цена_2001_до_3000'),
        (3001, 5000, 'Цена_3001_до_5000'),
        (5001, 7000, 'Цена_5001_до_7000'),
        (7001, float('inf'), 'Цена_7001_и_выше')
    ]
    for min_price, max_price, filename in price_ranges:
        range_df = df_clean[(df_clean['Цена'] >= min_price) &
                            (df_clean['Цена'] <= max_price)]
        range_df = range_df.sort_values(by='Цена', ascending=True,
                                        ignore_index=True)
        range_df.to_csv(f'{csv_dir}/{filename}.csv', index=False, sep=',')
        save_to_excel(range_df, f'{excel_dir}/{filename}.xlsx')
        save_to_pdf(range_df, f'{pdf_dir}/{filename}.pdf')

# Блок 8: Вывод результатов
# Итоги обработки.
print("Обработка завершена:")
print(f"Всего строк в чистом файле: {len(df_clean)}")
print(f"Столбцы: {list(df_clean.columns)}")
print(f"Первые 5 строк чистого файла:\n{df_clean.head()}")
if 'Рейтинг' in df_clean.columns:
    print(f"Уникальных рейтингов: {len(unique_ratings)}")
if 'Цена' in df_clean.columns:
    print("Диапазоны цен созданы: 0-1000, 1001-2000, 2001-3000, "
          "3001-5000, 5001-7000, 7001+")
